import pytest
import tempfile
import os
from core.qa.reporter import QAReporter
from core.models import QAReportItem


def test_generate_report_xlsx():
    with tempfile.TemporaryDirectory() as tmpdir:
        report_path = os.path.join(tmpdir, "conversion_report.xlsx")
        
        items = [
            QAReportItem(
                slide_no=1,
                detected_type="Cover",
                applied_layout="Title slide simple",
                migration_mode="migration",
                font_replaced="",
                objects_moved=0,
                objects_deleted=2,
                overflow_risk="None",
                need_manual_review=False,
                comment="封面页迁移成功",
            ),
            QAReportItem(
                slide_no=2,
                detected_type="Content",
                applied_layout="One column",
                migration_mode="adaptation",
                font_replaced="Calibri→Poppins",
                objects_moved=3,
                objects_deleted=1,
                overflow_risk="Low",
                need_manual_review=True,
                comment="文本框可能溢出，建议检查",
            ),
        ]
        
        reporter = QAReporter()
        reporter.generate(items, report_path)
        
        assert os.path.exists(report_path)
        assert os.path.getsize(report_path) > 0

        from openpyxl import load_workbook
        wb = load_workbook(report_path)
        assert "转换报告" in wb.sheetnames
        ws = wb["转换报告"]
        assert ws.max_row == 3  # header + 2 items
        assert ws.max_column == 10


def test_generate_report_with_summary():
    with tempfile.TemporaryDirectory() as tmpdir:
        report_path = os.path.join(tmpdir, "report.xlsx")
        
        items = [
            QAReportItem(
                slide_no=1,
                detected_type="Cover",
                applied_layout="Title slide",
                migration_mode="migration",
            ),
        ]
        
        reporter = QAReporter()
        reporter.generate(items, report_path, summary={
            "总页数": 1,
            "迁移页数": 1,
            "适配页数": 0,
        })
        
        from openpyxl import load_workbook
        wb = load_workbook(report_path)
        assert "转换报告" in wb.sheetnames
        assert "汇总" in wb.sheetnames
