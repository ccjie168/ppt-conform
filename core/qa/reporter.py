from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.models import QAReportItem

HEADERS = [
    "Slide No.",
    "Detected Type",
    "Applied Layout",
    "Migration Mode",
    "Font Replaced",
    "Objects Moved",
    "Objects Deleted",
    "Overflow Risk",
    "Need Manual Review",
    "Comment",
]

HEADER_FILL = PatternFill(start_color="0A2F24", end_color="0A2F24", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Poppins")
WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


class QAReporter:
    def generate(self, items: list[QAReportItem], output_path: str, summary: dict | None = None):
        wb = Workbook()
        ws = wb.active
        ws.title = "转换报告"

        # Write headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Write data
        for row_idx, item in enumerate(items, 2):
            values = [
                item.slide_no,
                item.detected_type,
                item.applied_layout,
                item.migration_mode,
                item.font_replaced,
                item.objects_moved,
                item.objects_deleted,
                item.overflow_risk,
                "是" if item.need_manual_review else "否",
                item.comment,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if item.need_manual_review:
                    cell.fill = WARN_FILL

        # Column widths
        col_widths = [12, 20, 28, 16, 25, 15, 16, 14, 18, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Summary sheet
        if summary:
            ws2 = wb.create_sheet("汇总")
            ws2["A1"] = "转换汇总"
            ws2["A1"].font = Font(bold=True, size=14, color="0A2F24")
            row = 3
            for key, value in summary.items():
                ws2.cell(row=row, column=1, value=key).font = Font(bold=True)
                ws2.cell(row=row, column=2, value=value)
                row += 1
            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 50

        wb.save(output_path)
